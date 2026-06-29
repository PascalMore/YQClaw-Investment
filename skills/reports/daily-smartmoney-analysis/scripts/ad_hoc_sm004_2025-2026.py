#!/usr/bin/env python3
"""Ad-hoc Smart Money push to Telegram (multi-day / single-product / custom threshold).

This is a TEMPLATE. Copy to your own script and modify.

Usage:
    PYTHONPATH=/home/pascal/workspace/yquant-investment:skills/data/data-pipeline/scripts \
        /home/pascal/workspace/yquant-investment/.venv/bin/python \
        ad_hoc_export.py

What it does:
    1. Query N days of portfolio_position for one product, take Top K by holding_ratio
    2. Query same N days of portfolio_trade for one product (no threshold)
    3. Generate Excel with two sheets (持仓 / 交易), date + product_code columns included
    4. Send via Telegram bot to DAILY_HAPPY_GROUP_CHAT_ID (or override)
    5. Does NOT update .last_sent marker (this is ad-hoc, not the daily cron)

PITFALLS to remember when adapting:
    - holding_ratio is stored as decimal (0.05 not 5%); display as f"{x*100:.2f}%"
    - change_amount sign convention is INVERTED in source data; trust the `direction` field
    - Two-day query needs $gte/$lte on the business date field, not $in
    - sort by holding_ratio DESC before limit(N) for true top-K
    - DO NOT hardcode a MongoDB URI — credentials live in skills/.env and
      the production script uses `PortfolioMongoLoader._db()`. Hardcoding
      is a security smell AND breaks the moment the password changes.
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

# Reuse the production script's helpers + the project loader
sys.path.insert(0, "/home/pascal/workspace/yquant-investment/skills/reports/daily-smartmoney-analysis/scripts")
sys.path.insert(0, "/home/pascal/workspace/yquant-investment/skills/data/data-pipeline/scripts")

from daily_export_report import (  # noqa: E402
    load_config,
    send_telegram_file,
    get_product_info_map,
    format_percentage,
    DEFAULT_ENV_PATH,
)
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
import pandas as pd  # noqa: E402
from loaders.mongodb_loader import PortfolioMongoLoader  # noqa: E402


def fetch(db, product_code: str, dates: list[str], top_k: int = 11) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fetch positions (Top K) and trades for a product over a date range.

    Returns (positions_df, trades_df).
    """
    product_info = get_product_info_map(db)
    info = product_info.get(product_code, {})

    # ---- nav snapshot per date (join key: position_date == nav_date) ----
    nav_cur = db["portfolio_nav"].find({
        "product_code": product_code,
        "nav_date": {"$gte": dates[0], "$lte": dates[-1]},
    })
    nav_by_date = {r["nav_date"]: (r.get("nav"), r.get("aum")) for r in nav_cur}

    # ---- positions: top K by holding_ratio, across all dates ----
    pos_cur = db["portfolio_position"].find({
        "product_code": product_code,
        "position_date": {"$gte": dates[0], "$lte": dates[-1]},
    })
    pos_rows = []
    for r in pos_cur:
        nav, aum = nav_by_date.get(r["position_date"], (None, None))
        pos_rows.append({
            "日期": r["position_date"],
            "产品代码": product_code,
            "产品名称": info.get("product_name", ""),
            "Wind代码": r["asset_wind_code"],
            "资产名称": r["asset_name"],
            "持仓比例": r["holding_ratio"],
            "数量(股)": int(r.get("shares", 0)),
            "市值(元)": int(r.get("market_value", 0)),
            "净值": nav,
            "规模(元)": int(aum) if aum is not None else None,
        })
    pos_df = pd.DataFrame(pos_rows)
    if not pos_df.empty and top_k > 0:
        # Top K per date
        pos_df = (pos_df
                  .sort_values(["日期", "持仓比例"], ascending=[True, False])
                  .groupby("日期", as_index=False)
                  .head(top_k)
                  .reset_index(drop=True))
    if not pos_df.empty:
        pos_df = pos_df.sort_values(["日期", "持仓比例"], ascending=[True, False]).reset_index(drop=True)
        pos_df = format_percentage(pos_df, ["持仓比例"])

    # ---- trades: all rows for the product on given dates ----
    trd_cur = db["portfolio_trade"].find({
        "product_code": product_code,
        "trade_date": {"$gte": dates[0], "$lte": dates[-1]},
    }).sort([("trade_date", 1), ("direction", 1)])
    trd_rows = []
    for r in trd_cur:
        trd_rows.append({
            "日期": r["trade_date"],
            "产品代码": product_code,
            "产品名称": info.get("product_name", ""),
            "Wind代码": r["asset_wind_code"],
            "资产名称": r["asset_name"],
            "方向": r["direction"],
            "变动比例": r["change_ratio"],
            "变动金额(元)": int(r.get("change_amount", 0)),
        })
    trd_df = pd.DataFrame(trd_rows)
    if not trd_df.empty:
        trd_df = format_percentage(trd_df, ["变动比例"])

    return pos_df, trd_df


def write_excel(pos_df: pd.DataFrame, trd_df: pd.DataFrame, output_path: str, top_k: int) -> None:
    wb = Workbook()

    for sheet_name, df in [("持仓", pos_df), ("交易", trd_df)]:
        if sheet_name == "持仓":
            ws = wb.active
            ws.title = "持仓" if (df.empty or top_k == 0) else f"持仓Top{top_k}"
        else:
            ws = wb.create_sheet(sheet_name)
        if df.empty:
            ws.cell(row=1, column=1, value="(无数据)")
            continue
        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True)
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        # Format percentage columns
        for ci, h in enumerate(headers, 1):
            if h in ("持仓比例", "变动比例"):
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=ci)
                    if isinstance(cell.value, str) and cell.value.endswith("%"):
                        cell.number_format = "0.00%"
            if h == "净值":
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=ci)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0000"

    wb.save(output_path)
    print(f"✅ Excel: {output_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--product", default="SM004", help="product_code, e.g. SM004")
    ap.add_argument("--start", default="2025-06-30", help="YYYY-MM-DD")
    ap.add_argument("--end", default="2026-06-25", help="YYYY-MM-DD")
    ap.add_argument("--top-k", type=int, default=0, help="Top K by holding_ratio per date; 0 = all positions (no limit)")
    ap.add_argument("--caption", default=None, help="Telegram caption override")
    ap.add_argument("--output", default=None, help="Excel output path (default /tmp/...)")
    ap.add_argument("--dry-run", action="store_true", help="Don't send, just generate Excel")
    ap.add_argument(
        "--to-personal",
        action="store_true",
        help="Send to Pascal personal chat (TELEGRAM_CHAT_ID) for preview confirmation; "
             "default is to refuse sending (dry-run only).",
    )
    args = ap.parse_args()

    dates = pd.date_range(args.start, args.end).strftime("%Y-%m-%d").tolist()
    output_path = args.output or f"/tmp/数据_{args.product}_{args.start}_{args.end}.xlsx"
    caption = args.caption or f"📊 {args.product} 持仓交易 {args.start}~{args.end} (确认预览)"

    # Load env for telegram config
    config = load_config(DEFAULT_ENV_PATH)
    bot_token = config.get("telegram_bot_token", "")
    personal_chat_id = config.get("telegram_chat_id", "")
    group_chat_id = config.get("daily_happy_chat_id", "")

    if not bot_token:
        print("❌ TELEGRAM_BOT_TOKEN not set")
        sys.exit(1)

    if args.to_personal:
        if not personal_chat_id:
            print("❌ TELEGRAM_CHAT_ID not set — refusing to fall back to Pascal's personal chat_id")
            sys.exit(1)
        chat_id = personal_chat_id
    else:
        # CRITICAL: refuse to fall back to Pascal's personal chat_id.
        # That's what daily_export_report.py silently does — and it has caused
        # production data leakage in the past. Force the operator to set the env var.
        if not group_chat_id:
            print("❌ DAILY_HAPPY_GROUP_CHAT_ID not set — refusing to fall back to Pascal's personal chat_id")
            sys.exit(1)
        chat_id = group_chat_id

    # Query via PortfolioMongoLoader (credentials from skills/.env)
    db = PortfolioMongoLoader()._db()
    pos_df, trd_df = fetch(db, args.product, dates, top_k=args.top_k)

    print(f"持仓: {len(pos_df)} 条, 交易: {len(trd_df)} 条")

    # Excel
    write_excel(pos_df, trd_df, output_path, args.top_k)

    # Send
    if args.dry_run:
        target = "personal" if args.to_personal else "group"
        print(f"🧪 DRY RUN — would send to {target} {chat_id}: {caption}")
        return

    success = send_telegram_file(bot_token, chat_id, output_path, caption)
    if not success:
        sys.exit(1)
    target = "personal" if args.to_personal else "group"
    print(f"✅ Sent to {target}: {caption}")


if __name__ == "__main__":
    main()
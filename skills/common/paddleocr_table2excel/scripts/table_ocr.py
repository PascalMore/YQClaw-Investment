#!/usr/bin/env python3
"""
PaddleOCR Table-to-Excel Pipeline
==================================
接收图片（Excel截图、手机照片、屏幕截图），
通过 PaddleOCR 逐行锚点扫描，输出标准 .xlsx 文件。

持仓表结构（以实际截图为准）：
  - 产品代码：SM001/SM881/SMO01/JS-001 等（系列标识，非合同号）
  - 持仓比例：4.62%（百分数字符串，x≈940）
  - Wind代码：300776.SZ（x≈670）
  - 资产名称：中文股票名（x≈806）
  - 数量：5-6位整数（x≈1036）
  - 市值：7位+整数（x≈1146）
  - 净值/份额：1.2849（x≈446）
  - 实际行高 ~25px，相邻 JS-001 之间隔 25±2px

依赖：PaddlePaddle 3.0.0 + PaddleOCR 2.7.3（全程离线）
"""
import sys as _sys
from pathlib import Path as _Path
_venv = _Path(__file__).parent / ".venv" / "lib" / "python3.10" / "site-packages"
if str(_venv) not in _sys.path:
    _sys.path.insert(0, str(_venv))

import cv2, re, warnings
from openpyxl import Workbook
warnings.filterwarnings("ignore")

from paddleocr import PaddleOCR


PRODUCT_CODES = {
    "SM001", "SM002", "SM003", "SM881", "SM882", "SM883",
    "SMO01", "SMO02", "SMO03", "SMOO",
    "JS-001", "JS-002", "JS-003", "JS-004",
    "ZO-001", "ZO-002", "ZO-003",
    "US-001", "US-002",
}

DATE_RE = re.compile(r"\b(\d{4}-\d{2}-\d{2})\b")
PCT_RE  = re.compile(r"\b(\d+(?:\.\d+)?)%\b")
SZ_RE   = re.compile(r"\b(\d{6}\.[A-Z]{2,4})\b")
HEADER  = ["截止日期","产品名称","产品代码","Wind代码","资产名称",
          "持仓比例","数量","市值（本币）","最新净值","最新份额","最新规模"]


def clean(s):
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s


class TablePreprocessor:
    @staticmethod
    def denoise(img):
        return cv2.fastNlMeansDenoisingColored(img, None, 8, 8, 7, 21)

    @staticmethod
    def enhance(img):
        lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        l = cv2.createCLAHE(2.0, (8, 8)).apply(l)
        return cv2.cvtColor(cv2.merge([l, a, b]), cv2.COLOR_LAB2BGR)

    def process(self, img):
        return self.enhance(self.denoise(img))


class Pipeline:
    def __init__(self):
        self.preprocessor = TablePreprocessor()
        self.ocr = PaddleOCR(lang="ch")

    def run(self, image_path, output_path=None):
        img = cv2.imread(str(image_path))
        if img is None:
            raise FileNotFoundError(f"无法读取图片: {image_path}")

        img = self.preprocessor.process(img)
        tmp_crop = str(image_path) + ".crop.jpg"
        cv2.imwrite(tmp_crop, img)

        result = self.ocr.ocr(img)
        if not result or not result[0]:
            return {"rows": 0, "columns": len(HEADER), "output": output_path, "header": HEADER}

        lines = result[0]

        # ── 1. 构建 token 列表 ───────────────────────────────────
        tokens = []   # (y, x, text, line_idx)
        for idx, line in enumerate(lines):
            pts = line[0]
            y = (pts[0][1] + pts[1][1] + pts[2][1] + pts[3][1]) / 4
            x = (pts[0][0] + pts[1][0] + pts[2][0] + pts[3][0]) / 4
            text = clean(line[1][0])
            if text:
                tokens.append((y, x, text, idx))

        tokens.sort(key=lambda t: t[0])

        HEADER_KW = {
            "产代", "产品", "最新", "规模", "业绩前五持仓", "云文档", "!!!",
            "Wind代码", "资产名称", "持仓比例", "数量", "市值（本币）",
            "最新净值", "最新份额", "最新规模", "截止日期",
        }

        # ── 2. 找到所有 JS-001 等锚点的 Y 坐标 ─────────────────────
        code_y_map = {}   # code_val -> list of y positions
        for y, x, text, _ in tokens:
            if text in PRODUCT_CODES:
                if text not in code_y_map:
                    code_y_map[text] = []
                code_y_map[text].append(y)

        # ── 3. 对每个锚点单独找同行 tokens（±15px 容差）───────────
        results = []
        used_line_indices = set()

        for y_anchor, x_anchor, anchor_text in [
            (y, x, t) for t, ys in code_y_map.items() for y in ys
            for x, _ in [(next((x_i for y_i, x_i, t_i, i_i in tokens
                                 if y_i == y and t_i == t), None), None)]
            if (y, x, t) in [(y2, x2, t2) for y2, x2, t2, _ in tokens]
        ]:
            # Find tokens on same row (±15px) that haven't been claimed
            row_tokens = [
                (x, y, t, idx) for y, x, t, idx in tokens
                if abs(y - y_anchor) <= 15 and idx not in used_line_indices
            ]
            if not row_tokens:
                continue

            used_line_indices.update(idx for _, _, _, idx in row_tokens)

            row_tokens.sort(key=lambda r: r[0])
            all_texts   = [t for _, _, t, _ in row_tokens]
            combined    = " ".join(all_texts)

            date_m = DATE_RE.search(combined)
            if not date_m:
                continue
            date_str = date_m.group(1)

            code_x = next((x for x, y, t, idx in row_tokens if t == anchor_text), None)
            if code_x is None:
                continue

            left_tokens  = [(x, t) for x, y, t, idx in row_tokens if x < code_x - 5]
            right_tokens = [(x, t) for x, y, t, idx in row_tokens if x > code_x + 5]

            # 产品名称
            product_name = ""
            for x, t in reversed(left_tokens):
                if re.search(r"[\u4e00-\u9fff]", t) and t not in HEADER_KW:
                    if re.match(r"^\d", t):
                        continue
                    product_name = t
                    break

            right_texts   = [t for _, t in right_tokens]
            data_tokens   = [t for t in right_texts if t not in HEADER_KW]
            data_combined = " ".join(right_texts)

            wind  = next((t for t in data_tokens if SZ_RE.match(t)), "")
            asset = next((t for t in data_tokens
                         if re.search(r"[\u4e00-\u9fff]", t) and t not in PRODUCT_CODES), "")
            pct_m = PCT_RE.search(data_combined)
            ratio = pct_m.group(0) if pct_m else ""

            wind_digits = re.search(r"\d{6}", wind).group() if wind else ""

            qty = ""
            for tok in data_tokens:
                if tok in (wind_digits, qty):
                    continue
                if re.match(r"^\d{4,6}$", tok):
                    qty = tok
                    break

            mkt = ""
            for tok in data_tokens:
                if tok in (wind_digits, qty):
                    continue
                if re.match(r"^\d{7,}$", tok):
                    mkt = tok
                    break

            DECIMAL_RE = re.compile(r"\b\d+\.\d+\b")
            pct_pos = data_combined.find("%")
            net = share = ""
            for d in DECIMAL_RE.findall(data_combined):
                if data_combined.find(d) > pct_pos:
                    if not net:
                        net = d
                    elif not share:
                        share = d

            results.append([
                date_str, product_name, anchor_text,
                wind, asset,
                ratio, qty, mkt,
                net, share, ""
            ])

        # ── 4. 去重 ───────────────────────────────────────────────
        seen = set()
        unique = []
        for row in results:
            key = (row[0], row[2], row[4])
            if key not in seen:
                seen.add(key)
                unique.append(row)

        # ── 5. 输出 Excel ─────────────────────────────────────────
        if output_path is None:
            output_path = str(_Path(image_path).with_suffix(".xlsx"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c_idx, col in enumerate(HEADER, 1):
            ws.cell(row=1, column=c_idx, value=col)
        for r_idx, row in enumerate(unique, 2):
            for c_idx, val in enumerate(row, 1):
                if val == "" or val is None:
                    continue
                try:
                    if isinstance(val, str) and "%" in val:
                        ws.cell(row=r_idx, column=c_idx, value=val)
                    elif "." in str(val):
                        ws.cell(row=r_idx, column=c_idx, value=float(val))
                    elif str(val).isdigit():
                        ws.cell(row=r_idx, column=c_idx, value=int(float(val)))
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=str(val))
                except (ValueError, AttributeError):
                    ws.cell(row=r_idx, column=c_idx, value=str(val))
        wb.save(str(output_path))

        if _Path(tmp_crop).exists():
            _Path(tmp_crop).unlink()

        return {
            "rows": len(unique),
            "columns": len(HEADER),
            "output": output_path,
            "header": HEADER
        }


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="PaddleOCR 表格图片转 Excel")
    p.add_argument("--input", "-i", required=True)
    p.add_argument("--output", "-o")
    args = p.parse_args()
    result = Pipeline().run(args.input, args.output)
    print(f"✅ 识别完成：{result['rows']} 行 → {result['output']}")
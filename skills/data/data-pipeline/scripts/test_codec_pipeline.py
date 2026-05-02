#!/usr/bin/env python3
"""测试：Excel → JSON → Base64 → JSON 反解析循环验证"""
import json
import sys
sys.path.insert(0, "scripts")

import openpyxl
from serializers.base64_codec import Base64Codec, encode_json, decode_base64

# ========== Step 1: Excel → JSON ==========
print("=" * 60)
print("测试 1: Excel → JSON → Base64")
print("=" * 60)

wb = openpyxl.load_workbook("examples/示例_5产品_250持仓.xlsx")
ws = wb.active

headers = [cell.value for cell in ws[1]]
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    record = dict(zip(headers, row))
    records.append(record)

json_data = {
    "metadata": {
        "source": "示例_5产品_250持仓.xlsx",
        "total_records": len(records),
        "columns": headers,
    },
    "data": records,
}

json_str = json.dumps(json_data, ensure_ascii=False, indent=2, default=str)
print(f"\n[Excel] 行数: {len(records)}")
print(f"[Excel] 列数: {len(headers)}")
print(f"[JSON] 字符数: {len(json_str)}")
print(f"\n[JSON] 结构预览:")
print(json_str[:500], "...\n")

# ========== Step 2: JSON → Base64 (gzip) ==========
codec_gzip = Base64Codec(compress=True)
b64_gzip = codec_gzip.encode(json_data)

print(f"[Base64 Gzip] 长度: {len(b64_gzip)}")
print(f"\n[Base64 Gzip] 前80字符:\n{b64_gzip[:80]}...")

# ========== Step 3: Base64 → JSON 反解析 ==========
print("\n" + "=" * 60)
print("测试 2: Base64 → JSON 反解析验证")
print("=" * 60)

decoded_gzip = codec_gzip.decode(b64_gzip)
match_gzip = (decoded_gzip == json_data)
print(f"\n[Gzip] 反解析成功: {match_gzip}")
print(f"[Gzip] 数据一致性: {decoded_gzip['metadata']['total_records']} 条记录")

# ========== 分析总结 ==========
print("\n" + "=" * 60)
print("测试结果分析")
print("=" * 60)

analysis = {
    "Excel数据量": f"{len(records)} 行 × {len(headers)} 列",
    "JSON字符数": len(json_str),
    "Base64_Gzip长度": len(b64_gzip),
    "Plain循环一致性": "N/A (已移除)",
    "Gzip循环一致性": "✅ 通过" if match_gzip else "❌ 失败",
}
for k, v in analysis.items():
    print(f"  {k}: {v}")

# 写入输出文件
output_b64 = "examples/output_base64.txt"
output_json = "examples/output_decoded.json"

with open(output_b64, "w") as f:
    f.write(b64_gzip)

with open(output_json, "w") as f:
    json.dump(decoded_gzip, f, ensure_ascii=False, indent=2, default=str)

print(f"\n输出文件已保存:")
print(f"  Base64: {output_b64}")
print(f"  JSON:   {output_json}")
